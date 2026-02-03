# excel_writer.py
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from sqlalchemy.orm import Session
from models import Assignment, Employee, ShiftDefinition, Workplace
from datetime import timedelta


def create_excel_report_from_db(session: Session, workplace_id: int, start_date):
    """
    Generates a visual Excel schedule based on assignments stored in the database.
    """
    # Fetch workplace info
    workplace = session.query(Workplace).get(workplace_id)
    employees = [e for e in workplace.employees if e.is_active]
    shifts = workplace.shifts

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Schedule"
    ws.sheet_view.rightToLeft = True  # Optimized for Hebrew users

    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    # Build Header Row (Dates)
    ws.cell(row=1, column=1).value = "Shift / Day"
    for d in range(7):
        current_date = start_date + timedelta(days=d)
        cell = ws.cell(row=1, column=d + 2)
        # Displaying Day Name and Date
        day_name = ["ראשון", "שני", "שלישי", "רביעי", "חמישי", "שישי", "שבת"][d]
        cell.value = f"{day_name}\n{current_date.strftime('%d/%m')}"
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    # Fill Assignments
    current_row = 2
    for s_def in shifts:
        # Each shift might have multiple slots (num_staff)
        for slot in range(s_def.num_staff):
            ws.cell(row=current_row, column=1).value = f"{s_def.shift_name} ({slot + 1})"
            ws.cell(row=current_row, column=1).alignment = center_align
            ws.cell(row=current_row, column=1).border = thin_border

            for d in range(7):
                target_date = start_date + timedelta(days=d)
                # Fetch assignment for this specific date, shift, and slot
                # We order by ID to maintain consistent slotting
                assign = session.query(Assignment).filter(
                    Assignment.workplace_id == workplace_id,
                    Assignment.shift_id == s_def.id,
                    Assignment.date == target_date
                ).all()

                cell = ws.cell(row=current_row, column=d + 2)
                cell.border = thin_border
                cell.alignment = center_align

                if len(assign) > slot:
                    emp = assign[slot].employee
                    cell.value = emp.name
                    if emp.color:
                        cell.fill = PatternFill(start_color=emp.color, end_color=emp.color, fill_type="solid")

            current_row += 1
        current_row += 1  # Add a gap between different shift types

    # Save file
    filename = f"schedule_{start_date.strftime('%Y%m%d')}.xlsx"
    wb.save(filename)
    print(f"✅ Visual Excel report saved as: {filename}")